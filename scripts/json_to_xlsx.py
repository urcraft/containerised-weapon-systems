#!/usr/bin/env python3
"""Generate the editable spreadsheet (data/systems.xlsx) FROM data/systems.json.

This is the inverse of build_data.py. It exists so the spreadsheet can be
regenerated from JSON if it is ever lost or the two drift apart. Normal editing
happens directly in the .xlsx; you should not usually need to run this.

Run with:  python scripts/json_to_xlsx.py
"""
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
FLAT_FIELDS = [
    "id", "name", "aka", "builder", "country", "sortDate", "dateLabel",
    "dateBasis", "category", "concealment", "summary", "confidence", "notes",
]
PHOTO_FIELDS = ["file", "credit", "license", "source"]
MAX_SOURCES = 3

HEADER = (
    FLAT_FIELDS
    + [f"photo_{k}" for k in PHOTO_FIELDS]
    + [f"source{n}_{p}" for n in range(1, MAX_SOURCES + 1) for p in ("url", "title")]
)

# Roughly sensible column widths for hand-editing.
WIDTHS = {
    "id": 22, "name": 34, "aka": 30, "builder": 34, "country": 16,
    "sortDate": 12, "dateLabel": 12, "dateBasis": 30, "category": 20,
    "concealment": 22, "summary": 70, "confidence": 12, "notes": 40,
}


def main():
    data = json.loads((ROOT / "data" / "systems.json").read_text(encoding="utf-8"))
    wb = Workbook()
    ws = wb.active
    ws.title = "systems"
    ws.append(HEADER)

    for rec in data:
        row = [rec.get(f, "") for f in FLAT_FIELDS]
        photo = rec.get("photo") or {}
        row += [photo.get(k, "") for k in PHOTO_FIELDS]
        sources = rec.get("sources") or []
        for n in range(MAX_SOURCES):
            s = sources[n] if n < len(sources) else {}
            row += [s.get("url", ""), s.get("title", "")]
        ws.append(row)

    # Header styling + freeze the header row.
    bold = Font(bold=True)
    for col, name in enumerate(HEADER, start=1):
        c = ws.cell(row=1, column=col)
        c.font = bold
        c.alignment = Alignment(vertical="center")
        letter = c.column_letter
        ws.column_dimensions[letter].width = WIDTHS.get(name, 26)
    ws.freeze_panes = "A2"

    # Turn the range into a filterable Excel table.
    last_col = ws.cell(row=1, column=len(HEADER)).column_letter
    ref = f"A1:{last_col}{ws.max_row}"
    table = Table(displayName="Systems", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)

    out = ROOT / "data" / "systems.xlsx"
    wb.save(out)
    print(f"Wrote {len(data)} systems -> {out}")


if __name__ == "__main__":
    main()
