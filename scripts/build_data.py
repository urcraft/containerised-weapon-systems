#!/usr/bin/env python3
"""Convert the editable spreadsheet (data/systems.xlsx) into data/systems.json.

The spreadsheet is the single source of truth for the card data. This script
flattens it back into the nested JSON shape that index.html fetches at runtime:

  - `photo`   is rebuilt from the photo_* columns
  - `sources` is rebuilt from the source{N}_url / source{N}_title columns

Run locally with:  python scripts/build_data.py
GitHub Actions runs the same script at deploy time (see .github/workflows/deploy.yml).
"""
import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "systems.xlsx"

# Flat card fields, in the order they should appear in each JSON record.
FLAT_FIELDS = [
    "id", "name", "aka", "builder", "country", "sortDate", "dateLabel",
    "dateBasis", "category", "concealment", "summary", "confidence", "notes",
]
PHOTO_FIELDS = ["file", "credit", "license", "source"]  # -> photo_file, photo_credit, ...
MAX_SOURCES = 3  # source1_url/title .. source3_url/title


def cell(value):
    """Normalise a spreadsheet cell to a trimmed string ('' for blanks)."""
    if value is None:
        return ""
    return str(value).strip()


def rows_to_records(rows):
    header, *body = rows
    header = [cell(h) for h in header]
    idx = {name: i for i, name in enumerate(header)}

    def get(row, col):
        return cell(row[idx[col]]) if col in idx and idx[col] < len(row) else ""

    records = []
    for row in body:
        if not any(cell(c) for c in row):
            continue  # skip fully blank rows
        rec = {}
        for f in FLAT_FIELDS:
            rec[f] = get(row, f)

        photo = {k: get(row, f"photo_{k}") for k in PHOTO_FIELDS}
        if any(photo.values()):
            rec["photo"] = photo

        sources = []
        for n in range(1, MAX_SOURCES + 1):
            url = get(row, f"source{n}_url")
            title = get(row, f"source{n}_title")
            if url or title:
                sources.append({"url": url, "title": title})
        rec["sources"] = sources

        # Reorder so `sources`/`photo` sit where they did originally.
        ordered = {}
        for f in FLAT_FIELDS:
            ordered[f] = rec[f]
            if f == "summary":
                ordered["sources"] = rec["sources"]
        if "sources" not in ordered:
            ordered["sources"] = rec["sources"]
        if "photo" in rec:
            ordered["photo"] = rec["photo"]
        records.append(ordered)
    return records


def main():
    ap = argparse.ArgumentParser(description="Build systems.json from systems.xlsx")
    ap.add_argument("--xlsx", default=str(XLSX), help="input .xlsx path")
    ap.add_argument("--out", default=str(ROOT / "data" / "systems.json"),
                    help="output .json path")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit(f"{args.xlsx} has no rows")

    records = rows_to_records(rows)
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(records, indent=2, ensure_ascii=False) + "\n",
                   encoding="utf-8")
    print(f"Wrote {len(records)} systems -> {out}")


if __name__ == "__main__":
    main()
